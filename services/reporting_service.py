"""
Сервис для создания отчетов и их экспорта.
Использует pandas для обработки данных и openpyxl для Excel-экспорта.
"""

import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
import sqlite3
import os
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Импорты для визуализации
import matplotlib
matplotlib.use('Agg')  # Использовать non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure
import seaborn as sns

from services.base import BaseUtilityService
from utils.exceptions import (
    ValidationError, RequiredFieldError, InvalidFormatError,
    ValueOutOfRangeError, RecordNotFoundError, BusinessLogicError
)
from utils.logger import get_logger

logger = get_logger('services.reporting')

# Настройка стилей для графиков
plt.style.use('seaborn-v0_8')
sns.set_palette("husl")


class ReportingService(BaseUtilityService):
    """
    Сервис для создания отчетов и их экспорта.
    """
    
    def __init__(self, db_connection: sqlite3.Connection):
        """
        Инициализация сервиса отчетности.
        
        Args:
            db_connection: Соединение с базой данных
        """
        super().__init__(db_connection)
        self.output_dir = "reports"
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """Создает директорию для отчетов если она не существует."""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def create_materials_report(self, 
                              supplier_id: Optional[int] = None,
                              grade_id: Optional[int] = None,
                              date_from: Optional[str] = None,
                              date_to: Optional[str] = None,
                              include_deleted: bool = False) -> pd.DataFrame:
        """
        Создает отчет по материалам с возможностью фильтрации.
        
        Args:
            supplier_id: ID поставщика для фильтрации
            grade_id: ID марки для фильтрации  
            date_from: Дата начала периода (YYYY-MM-DD)
            date_to: Дата окончания периода (YYYY-MM-DD)
            include_deleted: Включать удаленные материалы
            
        Returns:
            DataFrame с данными материалов
        """
        try:
            # Базовый запрос
            query = """
                SELECT 
                    m.id as 'ID',
                    m.arrival_date as 'Дата поступления',
                    s.name as 'Поставщик',
                    m.order_num as 'Номер заказа',
                    g.grade as 'Марка',
                    rt.type as 'Тип проката',
                    m.size as 'Размер',
                    m.cert_num as 'Номер сертификата',
                    m.cert_date as 'Дата сертификата',
                    m.batch as 'Партия',
                    m.heat_num as 'Номер плавки',
                    m.volume_length_mm as 'Длина (мм)',
                    m.volume_weight_kg as 'Масса (кг)',
                    CASE 
                        WHEN m.needs_lab = 1 THEN 'Да'
                        ELSE 'Нет'
                    END as 'Требует лабораторных исследований',
                    CASE 
                        WHEN m.to_delete = 1 THEN 'Помечен на удаление'
                        ELSE 'Активен'
                    END as 'Статус'
                FROM Materials m
                LEFT JOIN Suppliers s ON m.supplier_id = s.id
                LEFT JOIN Grades g ON m.grade_id = g.id
                LEFT JOIN RollingTypes rt ON m.rolling_type_id = rt.id
                WHERE 1=1
            """
            
            params = []
            
            # Добавляем условия фильтрации
            if not include_deleted:
                query += " AND m.to_delete = 0"
            
            if supplier_id:
                query += " AND m.supplier_id = ?"
                params.append(supplier_id)
            
            if grade_id:
                query += " AND m.grade_id = ?"
                params.append(grade_id)
            
            if date_from:
                query += " AND m.arrival_date >= ?"
                params.append(date_from)
            
            if date_to:
                query += " AND m.arrival_date <= ?"
                params.append(date_to)
            
            query += " ORDER BY m.arrival_date DESC"
            
            # Выполняем запрос и создаем DataFrame
            df = pd.read_sql_query(query, self.db_connection, params=params)
            
            logger.info(f"Создан отчет по материалам: {len(df)} записей")
            return df
            
        except Exception as e:
            self.handle_db_error(e, "создании отчета по материалам")
            raise
    
    def create_lab_requests_report(self,
                                 status_filter: Optional[str] = None,
                                 date_from: Optional[str] = None,
                                 date_to: Optional[str] = None) -> pd.DataFrame:
        """
        Создает отчет по заявкам лаборатории.
        
        Args:
            status_filter: Фильтр по статусу заявки
            date_from: Дата начала периода (YYYY-MM-DD)
            date_to: Дата окончания периода (YYYY-MM-DD)
            
        Returns:
            DataFrame с данными заявок
        """
        try:
            query = """
                SELECT 
                    lr.id as 'ID заявки',
                    lr.creation_date as 'Дата создания',
                    lr.request_number as 'Номер заявки',
                    m.id as 'ID материала',
                    s.name as 'Поставщик',
                    g.grade as 'Марка',
                    rt.type as 'Тип проката',
                    m.size as 'Размер',
                    lr.status as 'Статус',
                    CASE 
                        WHEN lr.archived = 1 THEN 'Да'
                        ELSE 'Нет'
                    END as 'Архивирована'
                FROM lab_requests lr
                LEFT JOIN Materials m ON lr.material_id = m.id
                LEFT JOIN Suppliers s ON m.supplier_id = s.id
                LEFT JOIN Grades g ON m.grade_id = g.id
                LEFT JOIN RollingTypes rt ON m.rolling_type_id = rt.id
                WHERE 1=1
            """
            
            params = []
            
            # Добавляем условия фильтрации
            if status_filter:
                query += " AND lr.status = ?"
                params.append(status_filter)
            
            if date_from:
                query += " AND lr.creation_date >= ?"
                params.append(date_from)
            
            if date_to:
                query += " AND lr.creation_date <= ?"
                params.append(date_to)
            
            query += " ORDER BY lr.creation_date DESC"
            
            # Выполняем запрос и создаем DataFrame
            df = pd.read_sql_query(query, self.db_connection, params=params)
            
            logger.info(f"Создан отчет по заявкам лаборатории: {len(df)} записей")
            return df
            
        except Exception as e:
            self.handle_db_error(e, "создании отчета по заявкам лаборатории")
            raise
    
    def create_statistics_report(self) -> Dict[str, Any]:
        """
        Создает сводный статистический отчет.
        
        Returns:
            Словарь со статистикой
        """
        try:
            stats = {}
            
            # Общая статистика материалов
            materials_query = """
                SELECT 
                    COUNT(*) as total_materials,
                    COUNT(CASE WHEN to_delete = 0 THEN 1 END) as active_materials,
                    COUNT(CASE WHEN to_delete = 1 THEN 1 END) as deleted_materials,
                    COUNT(CASE WHEN needs_lab = 1 THEN 1 END) as needs_lab_materials,
                    SUM(CASE WHEN to_delete = 0 THEN volume_weight_kg ELSE 0 END) as total_weight,
                    SUM(CASE WHEN to_delete = 0 THEN volume_length_mm ELSE 0 END) as total_length
                FROM Materials
            """
            
            materials_stats = pd.read_sql_query(materials_query, self.db_connection)
            stats['materials'] = materials_stats.iloc[0].to_dict()
            
            # Статистика по поставщикам
            suppliers_query = """
                SELECT 
                    s.name as supplier_name,
                    COUNT(m.id) as materials_count,
                    SUM(m.volume_weight_kg) as total_weight
                FROM Suppliers s
                LEFT JOIN Materials m ON s.id = m.supplier_id AND m.to_delete = 0
                GROUP BY s.id, s.name
                ORDER BY materials_count DESC
            """
            
            suppliers_stats = pd.read_sql_query(suppliers_query, self.db_connection)
            stats['suppliers'] = suppliers_stats.to_dict('records')
            
            # Статистика по маркам
            grades_query = """
                SELECT 
                    g.grade as grade_name,
                    COUNT(m.id) as materials_count,
                    SUM(m.volume_weight_kg) as total_weight
                FROM Grades g
                LEFT JOIN Materials m ON g.id = m.grade_id AND m.to_delete = 0
                GROUP BY g.id, g.grade
                ORDER BY materials_count DESC
            """
            
            grades_stats = pd.read_sql_query(grades_query, self.db_connection)
            stats['grades'] = grades_stats.to_dict('records')
            
            # Статистика по заявкам лаборатории
            lab_requests_query = """
                SELECT 
                    status,
                    COUNT(*) as count
                FROM lab_requests
                GROUP BY status
                ORDER BY count DESC
            """
            
            lab_requests_stats = pd.read_sql_query(lab_requests_query, self.db_connection)
            stats['lab_requests'] = lab_requests_stats.to_dict('records')
            
            logger.info("Создан статистический отчет")
            return stats
            
        except Exception as e:
            self.handle_db_error(e, "создании статистического отчета")
            raise
    
    def export_to_excel(self, data: pd.DataFrame, filename: str, sheet_name: str = "Отчет") -> str:
        """
        Экспортирует DataFrame в Excel с форматированием.
        
        Args:
            data: DataFrame для экспорта
            filename: Имя файла (без расширения)
            sheet_name: Название листа
            
        Returns:
            Путь к созданному файлу
        """
        try:
            # Создаем полный путь к файлу
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            full_filename = f"{filename}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, full_filename)
            
            # Создаем книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Добавляем данные
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)
            
            # Форматирование заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Стили для границ
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Применяем форматирование к заголовкам
            for col in range(1, len(data.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Форматирование данных
            data_alignment = Alignment(horizontal="left", vertical="center")
            
            for row in range(2, len(data) + 2):
                for col in range(1, len(data.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # Автоматическое изменение ширины столбцов
            for col in range(1, len(data.columns) + 1):
                column_letter = get_column_letter(col)
                max_length = max(
                    len(str(data.columns[col - 1])),
                    data.iloc[:, col - 1].astype(str).str.len().max() if len(data) > 0 else 0
                )
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Замораживаем первую строку
            ws.freeze_panes = "A2"
            
            # Сохраняем файл
            wb.save(filepath)
            
            logger.info(f"Отчет экспортирован в Excel: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка при экспорте в Excel: {e}")
            raise BusinessLogicError(
                f"Не удалось экспортировать отчет в Excel: {e}",
                suggestions=["Проверьте права доступа к папке отчетов"]
            )
    
    def export_statistics_to_excel(self, stats: Dict[str, Any], filename: str) -> str:
        """
        Экспортирует статистику в Excel с множественными листами.
        
        Args:
            stats: Словарь со статистикой
            filename: Имя файла (без расширения)
            
        Returns:
            Путь к созданному файлу
        """
        try:
            # Создаем полный путь к файлу
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            full_filename = f"{filename}_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, full_filename)
            
            # Создаем книгу Excel
            wb = Workbook()
            
            # Удаляем стандартный лист
            wb.remove(wb.active)
            
            # Создаем лист общей статистики
            ws_general = wb.create_sheet("Общая статистика")
            
            # Заполняем общую статистику
            general_data = [
                ["Показатель", "Значение"],
                ["Всего материалов", stats['materials']['total_materials']],
                ["Активных материалов", stats['materials']['active_materials']],
                ["Помечено на удаление", stats['materials']['deleted_materials']],
                ["Требует лабораторных исследований", stats['materials']['needs_lab_materials']],
                ["Общая масса (кг)", stats['materials']['total_weight']],
                ["Общая длина (мм)", stats['materials']['total_length']]
            ]
            
            for row_data in general_data:
                ws_general.append(row_data)
            
            # Форматируем общую статистику
            self._format_excel_sheet(ws_general, len(general_data), 2)
            
            # Создаем лист статистики поставщиков
            if stats['suppliers']:
                ws_suppliers = wb.create_sheet("Поставщики")
                suppliers_df = pd.DataFrame(stats['suppliers'])
                suppliers_df.columns = ['Поставщик', 'Количество материалов', 'Общая масса (кг)']
                
                for r in dataframe_to_rows(suppliers_df, index=False, header=True):
                    ws_suppliers.append(r)
                
                self._format_excel_sheet(ws_suppliers, len(suppliers_df) + 1, 3)
            
            # Создаем лист статистики марок
            if stats['grades']:
                ws_grades = wb.create_sheet("Марки")
                grades_df = pd.DataFrame(stats['grades'])
                grades_df.columns = ['Марка', 'Количество материалов', 'Общая масса (кг)']
                
                for r in dataframe_to_rows(grades_df, index=False, header=True):
                    ws_grades.append(r)
                
                self._format_excel_sheet(ws_grades, len(grades_df) + 1, 3)
            
            # Создаем лист статистики заявок
            if stats['lab_requests']:
                ws_lab = wb.create_sheet("Заявки лаборатории")
                lab_df = pd.DataFrame(stats['lab_requests'])
                lab_df.columns = ['Статус', 'Количество']
                
                for r in dataframe_to_rows(lab_df, index=False, header=True):
                    ws_lab.append(r)
                
                self._format_excel_sheet(ws_lab, len(lab_df) + 1, 2)
            
            # Сохраняем файл
            wb.save(filepath)
            
            logger.info(f"Статистика экспортирована в Excel: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка при экспорте статистики в Excel: {e}")
            raise BusinessLogicError(
                f"Не удалось экспортировать статистику в Excel: {e}",
                suggestions=["Проверьте права доступа к папке отчетов"]
            )
    
    def _format_excel_sheet(self, worksheet, row_count: int, col_count: int):
        """
        Применяет форматирование к листу Excel.
        
        Args:
            worksheet: Лист Excel
            row_count: Количество строк
            col_count: Количество столбцов
        """
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматирование заголовков
        for col in range(1, col_count + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Форматирование данных
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        for row in range(2, row_count + 1):
            for col in range(1, col_count + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Автоматическое изменение ширины столбцов
        for col in range(1, col_count + 1):
            column_letter = get_column_letter(col)
            worksheet.column_dimensions[column_letter].width = 25
        
        # Замораживаем первую строку
        worksheet.freeze_panes = "A2"
    
    def get_report_types(self) -> List[Dict[str, Any]]:
        """
        Возвращает список доступных типов отчетов.
        
        Returns:
            Список типов отчетов
        """
        return [
            {
                'id': 'materials',
                'name': 'Отчет по материалам',
                'description': 'Список материалов с возможностью фильтрации'
            },
            {
                'id': 'lab_requests',
                'name': 'Отчет по заявкам лаборатории',
                'description': 'Список заявок лаборатории с фильтрацией'
            },
            {
                'id': 'statistics',
                'name': 'Статистический отчет',
                'description': 'Общая статистика по материалам и заявкам'
            },
            {
                'id': 'quality_analysis',
                'name': 'Анализ качества по поставщикам',
                'description': 'Аналитический отчет по качеству продукции поставщиков'
            },
            {
                'id': 'defects_by_grade',
                'name': 'Статистика брака по маркам',
                'description': 'Анализ дефектов по маркам материалов'
            },
            {
                'id': 'supply_dynamics',
                'name': 'Динамика поставок',
                'description': 'Временной анализ поставок материалов'
            },
            {
                'id': 'overdue_tests',
                'name': 'Просроченные испытания',
                'description': 'Отчет о заявках с просроченными испытаниями'
            },
            {
                'id': 'dashboard',
                'name': 'Дашборд ключевых метрик',
                'description': 'Сводная панель с основными показателями'
            }
        ]
    
    # ===== АНАЛИТИЧЕСКИЕ ОТЧЕТЫ =====
    
    def create_quality_analysis_report(self, 
                                     date_from: Optional[str] = None,
                                     date_to: Optional[str] = None) -> pd.DataFrame:
        """
        Создает отчет по анализу качества по поставщикам.
        
        Args:
            date_from: Дата начала периода (YYYY-MM-DD)
            date_to: Дата окончания периода (YYYY-MM-DD)
            
        Returns:
            DataFrame с анализом качества по поставщикам
        """
        try:
            query = """
                SELECT 
                    s.name as 'Поставщик',
                    COUNT(m.id) as 'Всего материалов',
                    COUNT(CASE WHEN m.needs_lab = 1 THEN 1 END) as 'Требуют испытаний',
                    COUNT(d.id) as 'Количество дефектов',
                    ROUND(
                        CASE 
                            WHEN COUNT(m.id) > 0 
                            THEN (COUNT(d.id) * 100.0 / COUNT(m.id))
                            ELSE 0 
                        END, 2
                    ) as 'Процент брака (%)',
                    SUM(m.volume_weight_kg) as 'Общий вес (кг)',
                    GROUP_CONCAT(DISTINCT d.defect_type) as 'Типы дефектов'
                FROM Suppliers s
                LEFT JOIN Materials m ON s.id = m.supplier_id AND m.to_delete = 0
                LEFT JOIN defects d ON m.id = d.material_id AND d.to_delete = 0
                WHERE 1=1
            """
            
            params = []
            
            if date_from:
                query += " AND m.arrival_date >= ?"
                params.append(date_from)
            
            if date_to:
                query += " AND m.arrival_date <= ?"
                params.append(date_to)
            
            query += " GROUP BY s.id, s.name ORDER BY COUNT(m.id) DESC"
            
            df = pd.read_sql_query(query, self.db_connection, params=params)
            
            logger.info(f"Создан отчет по анализу качества поставщиков: {len(df)} записей")
            return df
            
        except Exception as e:
            self.handle_db_error(e, "создании отчета по анализу качества")
            raise
    
    def create_defects_by_grade_report(self,
                                     date_from: Optional[str] = None,
                                     date_to: Optional[str] = None) -> pd.DataFrame:
        """
        Создает отчет по статистике брака по маркам материалов.
        
        Args:
            date_from: Дата начала периода (YYYY-MM-DD)
            date_to: Дата окончания периода (YYYY-MM-DD)
            
        Returns:
            DataFrame со статистикой брака по маркам
        """
        try:
            query = """
                SELECT 
                    g.grade as 'Марка',
                    g.standard as 'Стандарт',
                    COUNT(m.id) as 'Всего материалов',
                    COUNT(d.id) as 'Количество дефектов',
                    ROUND(
                        CASE 
                            WHEN COUNT(m.id) > 0 
                            THEN (COUNT(d.id) * 100.0 / COUNT(m.id))
                            ELSE 0 
                        END, 2
                    ) as 'Процент брака (%)',
                    GROUP_CONCAT(DISTINCT d.defect_type) as 'Основные дефекты',
                    AVG(g.density) as 'Плотность',
                    SUM(m.volume_weight_kg) as 'Общий вес (кг)'
                FROM Grades g
                LEFT JOIN Materials m ON g.id = m.grade_id AND m.to_delete = 0
                LEFT JOIN defects d ON m.id = d.material_id AND d.to_delete = 0
                WHERE 1=1
            """
            
            params = []
            
            if date_from:
                query += " AND m.arrival_date >= ?"
                params.append(date_from)
            
            if date_to:
                query += " AND m.arrival_date <= ?"
                params.append(date_to)
            
            query += " GROUP BY g.id, g.grade, g.standard ORDER BY COUNT(d.id) DESC"
            
            df = pd.read_sql_query(query, self.db_connection, params=params)
            
            logger.info(f"Создан отчет по статистике брака по маркам: {len(df)} записей")
            return df
            
        except Exception as e:
            self.handle_db_error(e, "создании отчета по статистике брака")
            raise
    
    def create_supply_dynamics_report(self,
                                    date_from: Optional[str] = None,
                                    date_to: Optional[str] = None,
                                    group_by: str = 'month') -> pd.DataFrame:
        """
        Создает отчет по динамике поставок материалов.
        
        Args:
            date_from: Дата начала периода (YYYY-MM-DD)
            date_to: Дата окончания периода (YYYY-MM-DD)
            group_by: Группировка ('day', 'week', 'month', 'quarter')
            
        Returns:
            DataFrame с динамикой поставок
        """
        try:
            # Определяем формат группировки
            if group_by == 'day':
                date_format = "strftime('%Y-%m-%d', m.arrival_date)"
                period_name = 'День'
            elif group_by == 'week':
                date_format = "strftime('%Y-W%W', m.arrival_date)"
                period_name = 'Неделя'
            elif group_by == 'month':
                date_format = "strftime('%Y-%m', m.arrival_date)"
                period_name = 'Месяц'
            elif group_by == 'quarter':
                date_format = "strftime('%Y-Q', m.arrival_date) || CASE " \
                            "WHEN CAST(strftime('%m', m.arrival_date) AS INTEGER) <= 3 THEN '1' " \
                            "WHEN CAST(strftime('%m', m.arrival_date) AS INTEGER) <= 6 THEN '2' " \
                            "WHEN CAST(strftime('%m', m.arrival_date) AS INTEGER) <= 9 THEN '3' " \
                            "ELSE '4' END"
                period_name = 'Квартал'
            else:
                date_format = "strftime('%Y-%m', m.arrival_date)"
                period_name = 'Месяц'
            
            query = f"""
                SELECT 
                    {date_format} as '{period_name}',
                    COUNT(m.id) as 'Количество поставок',
                    COUNT(DISTINCT m.supplier_id) as 'Количество поставщиков',
                    SUM(m.volume_weight_kg) as 'Общий вес (кг)',
                    SUM(m.volume_length_mm) as 'Общая длина (мм)',
                    AVG(m.volume_weight_kg) as 'Средний вес поставки (кг)',
                    GROUP_CONCAT(DISTINCT s.name) as 'Поставщики'
                FROM Materials m
                LEFT JOIN Suppliers s ON m.supplier_id = s.id
                WHERE m.to_delete = 0
            """
            
            params = []
            
            if date_from:
                query += " AND m.arrival_date >= ?"
                params.append(date_from)
            
            if date_to:
                query += " AND m.arrival_date <= ?"
                params.append(date_to)
            
            query += f" GROUP BY {date_format} ORDER BY {date_format}"
            
            df = pd.read_sql_query(query, self.db_connection, params=params)
            
            logger.info(f"Создан отчет по динамике поставок: {len(df)} записей")
            return df
            
        except Exception as e:
            self.handle_db_error(e, "создании отчета по динамике поставок")
            raise
    
    def create_overdue_tests_report(self, overdue_days: int = 30) -> pd.DataFrame:
        """
        Создает отчет о просроченных испытаниях.
        
        Args:
            overdue_days: Количество дней для определения просрочки
            
        Returns:
            DataFrame с просроченными испытаниями
        """
        try:
            cutoff_date = (datetime.now() - timedelta(days=overdue_days)).strftime('%Y-%m-%d')
            
            query = """
                SELECT 
                    lr.request_number as 'Номер заявки',
                    lr.creation_date as 'Дата создания',
                    DATE(lr.creation_date, '+{} days') as 'Плановая дата завершения',
                    julianday('now') - julianday(lr.creation_date) as 'Дней просрочки',
                    lr.status as 'Статус',
                    m.id as 'ID материала',
                    s.name as 'Поставщик',
                    g.grade as 'Марка',
                    rt.type as 'Тип проката',
                    m.arrival_date as 'Дата поступления материала',
                    m.cert_num as 'Номер сертификата'
                FROM lab_requests lr
                LEFT JOIN Materials m ON lr.material_id = m.id
                LEFT JOIN Suppliers s ON m.supplier_id = s.id
                LEFT JOIN Grades g ON m.grade_id = g.id
                LEFT JOIN RollingTypes rt ON m.rolling_type_id = rt.id
                WHERE lr.creation_date <= ?
                AND lr.status NOT IN ('Завершена', 'Отменена', 'Архивирована')
                AND lr.archived = 0
                ORDER BY lr.creation_date ASC
            """.format(overdue_days)
            
            df = pd.read_sql_query(query, self.db_connection, params=[cutoff_date])
            
            logger.info(f"Создан отчет о просроченных испытаниях: {len(df)} записей")
            return df
            
        except Exception as e:
            self.handle_db_error(e, "создании отчета о просроченных испытаниях")
            raise
    
    def create_dashboard_data(self) -> Dict[str, Any]:
        """
        Создает данные для дашборда с ключевыми метриками.
        
        Returns:
            Словарь с данными для дашборда
        """
        try:
            dashboard_data = {}
            
            # Основные показатели
            main_metrics = self.db_connection.execute("""
                SELECT 
                    COUNT(*) as total_materials,
                    COUNT(CASE WHEN to_delete = 0 THEN 1 END) as active_materials,
                    COUNT(CASE WHEN needs_lab = 1 AND to_delete = 0 THEN 1 END) as need_testing,
                    SUM(CASE WHEN to_delete = 0 THEN volume_weight_kg ELSE 0 END) as total_weight,
                    COUNT(DISTINCT supplier_id) as suppliers_count
                FROM Materials
            """).fetchone()
            
            dashboard_data['main_metrics'] = {
                'total_materials': main_metrics[0],
                'active_materials': main_metrics[1],
                'need_testing': main_metrics[2],
                'total_weight': main_metrics[3] or 0,
                'suppliers_count': main_metrics[4]
            }
            
            # Статистика дефектов
            defects_stats = self.db_connection.execute("""
                SELECT 
                    COUNT(*) as total_defects,
                    COUNT(DISTINCT material_id) as affected_materials,
                    COUNT(DISTINCT defect_type) as defect_types_count
                FROM defects
                WHERE to_delete = 0
            """).fetchone()
            
            dashboard_data['defects_stats'] = {
                'total_defects': defects_stats[0],
                'affected_materials': defects_stats[1],
                'defect_types_count': defects_stats[2]
            }
            
            # Статистика лаборатории
            lab_stats = self.db_connection.execute("""
                SELECT 
                    COUNT(*) as total_requests,
                    COUNT(CASE WHEN status = 'Не отработана' THEN 1 END) as pending_requests,
                    COUNT(CASE WHEN status = 'Завершена' THEN 1 END) as completed_requests,
                    COUNT(CASE WHEN archived = 1 THEN 1 END) as archived_requests
                FROM lab_requests
            """).fetchone()
            
            dashboard_data['lab_stats'] = {
                'total_requests': lab_stats[0],
                'pending_requests': lab_stats[1],
                'completed_requests': lab_stats[2],
                'archived_requests': lab_stats[3]
            }
            
            # Топ поставщики
            top_suppliers = self.db_connection.execute("""
                SELECT 
                    s.name,
                    COUNT(m.id) as materials_count,
                    SUM(m.volume_weight_kg) as total_weight
                FROM Suppliers s
                LEFT JOIN Materials m ON s.id = m.supplier_id AND m.to_delete = 0
                GROUP BY s.id, s.name
                ORDER BY materials_count DESC
                LIMIT 5
            """).fetchall()
            
            dashboard_data['top_suppliers'] = [
                {'name': row[0], 'materials_count': row[1], 'total_weight': row[2] or 0}
                for row in top_suppliers
            ]
            
            # Последние поставки
            recent_supplies = self.db_connection.execute("""
                SELECT 
                    m.arrival_date,
                    s.name as supplier,
                    g.grade,
                    m.volume_weight_kg
                FROM Materials m
                LEFT JOIN Suppliers s ON m.supplier_id = s.id
                LEFT JOIN Grades g ON m.grade_id = g.id
                WHERE m.to_delete = 0
                ORDER BY m.arrival_date DESC
                LIMIT 10
            """).fetchall()
            
            dashboard_data['recent_supplies'] = [
                {
                    'date': row[0],
                    'supplier': row[1],
                    'grade': row[2],
                    'weight': row[3] or 0
                }
                for row in recent_supplies
            ]
            
            # Процент качества по поставщикам
            quality_by_suppliers = self.db_connection.execute("""
                SELECT 
                    s.name,
                    COUNT(m.id) as total_materials,
                    COUNT(d.id) as defects_count,
                    ROUND(
                        CASE 
                            WHEN COUNT(m.id) > 0 
                            THEN ((COUNT(m.id) - COUNT(d.id)) * 100.0 / COUNT(m.id))
                            ELSE 100 
                        END, 1
                    ) as quality_percentage
                FROM Suppliers s
                LEFT JOIN Materials m ON s.id = m.supplier_id AND m.to_delete = 0
                LEFT JOIN defects d ON m.id = d.material_id AND d.to_delete = 0
                GROUP BY s.id, s.name
                HAVING COUNT(m.id) > 0
                ORDER BY quality_percentage DESC
            """).fetchall()
            
            dashboard_data['quality_by_suppliers'] = [
                {
                    'supplier': row[0],
                    'total_materials': row[1],
                    'defects_count': row[2],
                    'quality_percentage': row[3]
                }
                for row in quality_by_suppliers
            ]
            
            logger.info("Создан дашборд с ключевыми метриками")
            return dashboard_data
            
        except Exception as e:
            self.handle_db_error(e, "создании дашборда")
            raise
    
    def create_chart_image(self, chart_type: str, data: Any, title: str = "", **kwargs) -> str:
        """
        Создает изображение графика и возвращает его в base64.
        
        Args:
            chart_type: Тип графика ('bar', 'line', 'pie', 'scatter')
            data: Данные для графика
            title: Заголовок графика
            **kwargs: Дополнительные параметры
            
        Returns:
            Строка в формате base64 с изображением графика
        """
        try:
            fig, ax = plt.subplots(figsize=(10, 6))
            fig.patch.set_facecolor('white')
            
            if chart_type == 'bar':
                if isinstance(data, pd.DataFrame):
                    x_col = kwargs.get('x_column', data.columns[0])
                    y_col = kwargs.get('y_column', data.columns[1])
                    ax.bar(data[x_col], data[y_col])
                    ax.set_xlabel(x_col)
                    ax.set_ylabel(y_col)
                
            elif chart_type == 'line':
                if isinstance(data, pd.DataFrame):
                    x_col = kwargs.get('x_column', data.columns[0])
                    y_col = kwargs.get('y_column', data.columns[1])
                    ax.plot(data[x_col], data[y_col], marker='o')
                    ax.set_xlabel(x_col)
                    ax.set_ylabel(y_col)
                    plt.xticks(rotation=45)
                
            elif chart_type == 'pie':
                if isinstance(data, pd.DataFrame):
                    labels_col = kwargs.get('labels_column', data.columns[0])
                    values_col = kwargs.get('values_column', data.columns[1])
                    ax.pie(data[values_col], labels=data[labels_col], autopct='%1.1f%%')
                
            elif chart_type == 'scatter':
                if isinstance(data, pd.DataFrame):
                    x_col = kwargs.get('x_column', data.columns[0])
                    y_col = kwargs.get('y_column', data.columns[1])
                    ax.scatter(data[x_col], data[y_col])
                    ax.set_xlabel(x_col)
                    ax.set_ylabel(y_col)
            
            ax.set_title(title)
            plt.tight_layout()
            
            # Сохраняем график в base64
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            buffer.seek(0)
            image_base64 = base64.b64encode(buffer.getvalue()).decode()
            plt.close(fig)
            
            return image_base64
            
        except Exception as e:
            logger.error(f"Ошибка при создании графика: {e}")
            plt.close('all')
            return ""
    
    def get_filter_options(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Возвращает доступные опции для фильтрации.
        
        Returns:
            Словарь с опциями фильтрации
        """
        try:
            # Получаем поставщиков
            suppliers_query = "SELECT id, name FROM Suppliers ORDER BY name"
            suppliers_df = pd.read_sql_query(suppliers_query, self.db_connection)
            suppliers = suppliers_df.to_dict('records')
            
            # Получаем марки
            grades_query = "SELECT id, grade FROM Grades ORDER BY grade"
            grades_df = pd.read_sql_query(grades_query, self.db_connection)
            grades = grades_df.to_dict('records')
            
            # Получаем статусы заявок
            statuses_query = "SELECT DISTINCT status FROM lab_requests ORDER BY status"
            statuses_df = pd.read_sql_query(statuses_query, self.db_connection)
            statuses = [{'id': row['status'], 'name': row['status']} for _, row in statuses_df.iterrows()]
            
            return {
                'suppliers': suppliers,
                'grades': grades,
                'lab_statuses': statuses
            }
            
        except Exception as e:
            logger.error(f"Ошибка при получении опций фильтрации: {e}")
            return {'suppliers': [], 'grades': [], 'lab_statuses': []}
    
    def validate_date_range(self, date_from: str, date_to: str) -> None:
        """
        Валидирует диапазон дат.
        
        Args:
            date_from: Дата начала
            date_to: Дата окончания
            
        Raises:
            ValidationError: При некорректном диапазоне дат
        """
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            
            if start_date > end_date:
                raise ValidationError(
                    "Дата начала не может быть больше даты окончания",
                    suggestions=["Проверьте правильность выбранных дат"]
                )
            
            # Проверяем, что диапазон не слишком большой (больше 5 лет)
            if (end_date - start_date).days > 365 * 5:
                raise ValidationError(
                    "Диапазон дат не может превышать 5 лет",
                    suggestions=["Выберите более короткий период"]
                )
                
        except ValueError:
            raise InvalidFormatError(
                "Некорректный формат даты. Используйте формат YYYY-MM-DD",
                suggestions=["Проверьте правильность ввода дат"]
            )
    
    def handle_db_error(self, error: Exception, operation: str) -> None:
        """
        Обработка ошибок базы данных.
        
        Args:
            error: Исключение
            operation: Описание операции
        """
        logger.error(f"Ошибка БД при {operation}: {error}")
        raise BusinessLogicError(
            f"Ошибка при {operation}: {error}",
            original_error=error,
            suggestions=["Проверьте соединение с базой данных"]
        ) 